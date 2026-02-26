#!/usr/bin/env python3
"""
convert_to_txt.py
-----------------
Traverses a root directory (recursively) and converts supported files to .txt,
saving them in an output directory with a flat filename that encodes the full
relative path:

    root_dir>subdir>...>original_filename.txt

Incremental mode (default): skips files whose content hash hasn't changed since
the last run. A manifest file (.conversion_manifest.json) is stored in the
output directory to track hashes.

Supported formats
-----------------
  .txt   → copied as-is (passthrough, change-tracked)
  .csv   → key: value | key: value  (one chunk per row, double-newline separated)
  .pdf   → pdftotext -layout; auto-falls back to tesseract OCR for image-based PDFs
  .docx  → pdftotext -layout via intermediate PDF conversion using LibreOffice
  .xlsx  → sheet-by-sheet table text via openpyxl (empty rows/cols stripped)

Usage
-----
    python convert_to_txt.py <root_dir> <output_dir>
    python convert_to_txt.py <root_dir> <output_dir> --force   # reprocess everything

Dependencies
------------
    pip install openpyxl pytesseract Pillow --break-system-packages
    apt-get install -y poppler-utils libreoffice tesseract-ocr
"""

import argparse
import csv
import io
import hashlib
import json
import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = {".csv", ".pdf", ".docx", ".xlsx", ".txt"}
MANIFEST_FILE = ".conversion_manifest.json"


def flat_output_name(root: Path, file_path: Path) -> str:
    rel = file_path.relative_to(root)
    parts = [root.name] + list(rel.parts)
    parts[-1] = Path(parts[-1]).stem + ".txt"
    return ">".join(parts)


def write_txt(output_dir: Path, filename: str, content: str) -> Path:
    out_path = output_dir / filename
    out_path.write_text(content, encoding="utf-8")
    return out_path


def file_hash(file_path: Path, chunk_size: int = 65536) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(chunk_size):
            h.update(chunk)
    return h.hexdigest()


def load_manifest(output_dir: Path) -> dict:
    manifest_path = output_dir / MANIFEST_FILE
    if manifest_path.exists():
        try:
            return json.loads(manifest_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_manifest(output_dir: Path, manifest: dict) -> None:
    manifest_path = output_dir / MANIFEST_FILE
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")


# ---------------------------------------------------------------------------
# Converters
# ---------------------------------------------------------------------------

def convert_csv(file_path: Path, separator: str = " | ") -> str:
    """Row-as-document: 'field_name: value | field_name: value' per row, double-newline separated."""
    with open(file_path, newline="", encoding="utf-8-sig") as f:
        return "\n\n".join(_csv_chunks_from_file(f, separator))


# Minimum non-whitespace characters pdftotext must return before we
# consider the PDF to have selectable text. Below this threshold we
# fall back to OCR.
PDF_TEXT_MIN_CHARS = 50


def _ocr_pdf(file_path: Path) -> str:
    """
    Fallback for image-based PDFs.
    Renders each page to an image with pdftoppm, then runs tesseract OCR.

    Requirements:
      apt-get install -y tesseract-ocr
      pip install pytesseract Pillow --break-system-packages
    """
    try:
        import pytesseract
        from PIL import Image
    except ImportError:
        raise ImportError(
            "OCR dependencies missing. Install with:\n"
            "  pip install pytesseract Pillow --break-system-packages\n"
            "  apt-get install -y tesseract-ocr"
        )

    with tempfile.TemporaryDirectory() as tmpdir:
        # Render PDF pages to PNG images at 300 DPI via pdftoppm
        result = subprocess.run(
            ["pdftoppm", "-r", "300", "-png", str(file_path), f"{tmpdir}/page"],
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            raise RuntimeError(f"pdftoppm failed: {result.stderr.strip()}")

        page_files = sorted(Path(tmpdir).glob("page-*.png"))
        if not page_files:
            raise RuntimeError("pdftoppm produced no output pages.")

        pages_text = []
        for page_img in page_files:
            text = pytesseract.image_to_string(Image.open(page_img))
            if text.strip():
                pages_text.append(text)

        return "\n\n".join(pages_text)


def convert_pdf(file_path: Path) -> str:
    """
    Extract text from a PDF using a two-step strategy:
      1. pdftotext -layout  — fast, accurate for text-based PDFs.
      2. If extracted text is below PDF_TEXT_MIN_CHARS (image-based / scanned),
         automatically fall back to tesseract OCR via pdftoppm + pytesseract.
    """
    result = subprocess.run(
        ["pdftotext", "-layout", str(file_path), "-"],
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(f"pdftotext failed: {result.stderr.strip()}")

    text = result.stdout
    if len(text.strip()) >= PDF_TEXT_MIN_CHARS:
        return text  # normal text-based PDF — done

    # Image-based PDF — fall back to OCR
    print(f"    ↳ No selectable text found, falling back to OCR...")
    return _ocr_pdf(file_path)


def convert_docx(file_path: Path) -> str:
    """Convert docx → PDF via LibreOffice headless, then extract with pdftotext -layout."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_docx = Path(tmpdir) / file_path.name
        shutil.copy2(file_path, tmp_docx)

        result = subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf", "--outdir", tmpdir, str(tmp_docx)],
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            raise RuntimeError(f"soffice conversion failed: {result.stderr.strip()}")

        pdf_path = Path(tmpdir) / (tmp_docx.stem + ".pdf")
        if not pdf_path.exists():
            raise FileNotFoundError(f"Expected PDF not found: {pdf_path}")

        return convert_pdf(pdf_path)


def convert_txt(file_path: Path) -> str:
    """Plain passthrough — read and return the file's text content as-is."""
    return file_path.read_text(encoding="utf-8", errors="replace")


def _csv_chunks_from_file(f, separator: str = " | ") -> list[str]:
    """
    Shared logic: parse a CSV file-like object and return a list of
    'field: value | field: value' strings, one per non-empty data row.
    Columns with an empty header are silently skipped.
    """
    chunks = []
    reader = csv.DictReader(f)
    for row in reader:
        line = separator.join(
            f"{k}: {v}"
            for k, v in row.items()
            if k and k.strip() and v and v.strip()
        )
        if line:
            chunks.append(line)
    return chunks


def convert_xlsx(file_path: Path, separator: str = " | ") -> str:
    """
    Convert each sheet to an in-memory CSV, then pass through the same
    _csv_chunks_from_file() function used by convert_csv — guaranteeing
    identical output format for both formats.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl --break-system-packages")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows(values_only=True))
        if not raw_rows:
            continue

        # Serialise sheet → in-memory CSV
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in raw_rows:
            writer.writerow(
                [str(cell).strip() if cell is not None else "" for cell in row]
            )
        buf.seek(0)

        # Reuse CSV chunk logic
        all_chunks.extend(_csv_chunks_from_file(buf, separator))

    return "\n\n".join(all_chunks)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

CONVERTERS = {
    ".csv":  convert_csv,
    ".pdf":  convert_pdf,
    ".docx": convert_docx,
    ".xlsx": convert_xlsx,
    ".txt":  convert_txt,
}


def process_file(
    file_path: Path,
    root: Path,
    output_dir: Path,
    manifest: dict,
    force: bool,
) -> tuple[bool, str | None]:
    ext = file_path.suffix.lower()
    converter = CONVERTERS.get(ext)
    if converter is None:
        return False, None

    manifest_key = str(file_path.relative_to(root))
    out_name = flat_output_name(root, file_path)
    out_path = output_dir / out_name

    current_hash = file_hash(file_path)
    stored_hash = manifest.get(manifest_key)

    # Skip if hash unchanged AND output file still exists
    if not force and stored_hash == current_hash and out_path.exists():
        print(f"  – Skipped (unchanged): {file_path.relative_to(root)}")
        return False, None

    reason = "forced" if force else ("new file" if stored_hash is None else "changed")
    print(f"  Converting [{reason}]: {file_path.relative_to(root)}  →  {out_name}")

    try:
        content = converter(file_path)
        write_txt(output_dir, out_name, content)
        manifest[manifest_key] = current_hash
        print(f"  ✓ Saved: {out_path}")
        return True, current_hash
    except Exception as exc:
        print(f"  ✗ FAILED ({file_path.name}): {exc}", file=sys.stderr)
        return False, None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Recursively convert CSV/PDF/DOCX/XLSX/TXT files to TXT for RAG pipelines."
    )
    parser.add_argument("root_dir", help="Root directory to traverse")
    parser.add_argument("output_dir", help="Directory where .txt files will be saved")
    parser.add_argument(
        "--force", action="store_true",
        help="Reprocess all files regardless of whether they have changed"
    )
    args = parser.parse_args()

    root = Path(args.root_dir).resolve()
    output_dir = Path(args.output_dir).resolve()

    if not root.is_dir():
        print(f"Error: '{root}' is not a directory.", file=sys.stderr)
        sys.exit(1)

    output_dir.mkdir(parents=True, exist_ok=True)

    manifest = load_manifest(output_dir)
    if args.force:
        print("--force specified: reprocessing all files.\n")

    files = [
        p for p in root.rglob("*")
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    ]

    if not files:
        print(f"No supported files found under '{root}'.")
        sys.exit(0)

    print(f"Found {len(files)} file(s). Checking for changes...\n")

    converted = skipped = failed = 0
    for file_path in sorted(files):
        processed, _ = process_file(file_path, root, output_dir, manifest, args.force)
        if processed:
            converted += 1
        else:
            manifest_key = str(file_path.relative_to(root))
            if manifest.get(manifest_key) and not args.force:
                skipped += 1
            elif not processed:
                failed += 1

    save_manifest(output_dir, manifest)

    print(f"\n{'─'*50}")
    print(f"  Converted : {converted}")
    print(f"  Skipped   : {skipped}  (unchanged)")
    print(f"  Failed    : {failed}")
    print(f"  Manifest  : {output_dir / MANIFEST_FILE}")
    print(f"  Output    : {output_dir}")


if __name__ == "__main__":
    main()